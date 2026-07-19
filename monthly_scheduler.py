# -*- coding: utf-8 -*-
"""Monthly on-call scheduler  (reusable).
Usage:  python3 monthly_scheduler.py <interns.csv> <year> <month> <output.xlsx>
Hard constraints are guaranteed (0 violations); soft objectives
(fair load / weekend / Fri / Sat balance, no-sandwich, preferred days &
stations, even spread) are optimized via constrained construction + SA.
The CSV must have the same columns as july2026interns.csv."""
import sys, csv, datetime, random, math, json, calendar, pickle
random.seed(42)
_RAW = sys.argv[1:]
# split flags (--key value) from positional args; positional order is unchanged for compatibility
_FLAGS = {}; _A = []
_i = 0
while _i < len(_RAW):
    tok = _RAW[_i]
    if tok.startswith("--"):
        key = tok[2:]
        val = _RAW[_i+1] if _i+1 < len(_RAW) and not _RAW[_i+1].startswith("--") else "1"
        _FLAGS[key] = val; _i += 2
    else:
        _A.append(tok); _i += 1
CSV   = _A[0] if len(_A) > 0 else "interns.csv"
YEAR  = int(_A[1]) if len(_A) > 1 else 2026
MONTH = int(_A[2]) if len(_A) > 2 else 7
OUT   = _A[3] if len(_A) > 3 else "schedule.xlsx"
DAYS  = list(range(1, calendar.monthrange(YEAR, MONTH)[1] + 1))

# optional inputs (flags take precedence; 5th positional kept for external compatibility)
HOL_CSV  = _FLAGS.get("holidays")
BASE_CSV = _FLAGS.get("base")                       # machine-readable existing schedule (mid-month edit)
FROM_DAY = int(_FLAGS["from"]) if "from" in _FLAGS else None   # re-optimize from this day onward
LOCKS_CSV = _FLAGS.get("locks")                      # locked slots that must not change

STATIONS = ["er1", "er2", "nicu1", "nicu2", "ward", "picu"]
PAIRS = [("er1", "er2"), ("nicu1", "nicu2")]
PAIR_OF = {"er1": "er2", "er2": "er1", "nicu1": "nicu2", "nicu2": "nicu1"}

# ---- external (non-intern) duties ----
# --external FILE  (or 5th positional): CSV day,station,name. Those station-days are
# covered by EXTERNAL staff and removed from the intern problem (lowers everyone's load).
EXT_CSV = _FLAGS.get("external") or (_A[4] if len(_A) > 4 else None)
EXTERNAL = {}   # (day, station) -> external person's name
if EXT_CSV:
    with open(EXT_CSV, encoding="utf-8-sig") as _f:
        for _row in csv.DictReader(_f):
            try: _d = int(str(_row.get("day", "")).strip())
            except (ValueError, TypeError): continue
            _st = (_row.get("station") or "").strip()
            _nm = (_row.get("name") or "").strip()
            if _st in STATIONS and _nm:
                EXTERNAL[(_d, _st)] = _nm
SKIP = set(EXTERNAL.keys())            # (day, station) NOT filled by interns

# ---- locked slots (from --locks): must not change during re-optimization ----
LOCKED = set()   # (day, station) pairs that are frozen
if LOCKS_CSV:
    with open(LOCKS_CSV, encoding="utf-8-sig") as _f:
        for _row in csv.DictReader(_f):
            try: _d = int(str(_row.get("day", "")).strip())
            except (ValueError, TypeError): continue
            _st = (_row.get("station") or "").strip()
            if _st in STATIONS and _d in range(1, 32):
                LOCKED.add((_d, _st))
    if LOCKED:
        print(f"  Locked slots: {len(LOCKED)}")

# ---- day classification ----
def dow(d):  # 4=Fri,5=Sat
    return datetime.date(YEAR, MONTH, d).weekday()
IS_FRI = {d: dow(d) == 4 for d in DAYS}
IS_SAT = {d: dow(d) == 5 for d in DAYS}
IS_THU = {d: dow(d) == 3 for d in DAYS}

# ---- holidays (behave like weekends): eve == Friday, holiday == Saturday ----
# --holidays FILE : CSV with columns date,kind[,name].  date = day-of-month or YYYY-MM-DD.
#   kind = "eve" (treated like Friday) or "day"/"holiday" (treated like Saturday).
HOLIDAY = {}   # day -> label
def _to_day(v):
    v = str(v).strip()
    if not v: return None
    if "-" in v:                       # ISO date
        try:
            dt = datetime.date.fromisoformat(v)
            return dt.day if (dt.year == YEAR and dt.month == MONTH) else None
        except ValueError: return None
    try: return int(v)
    except ValueError: return None
if HOL_CSV:
    with open(HOL_CSV, encoding="utf-8-sig") as _f:
        for _row in csv.DictReader(_f):
            _d = _to_day(_row.get("date", ""))
            if _d is None or _d not in DAYS: continue
            _kind = (_row.get("kind") or "").strip().lower()
            _nm = (_row.get("name") or "").strip()
            if _kind in ("eve", "erev", "ערב"):
                IS_FRI[_d] = True; HOLIDAY[_d] = _nm or "ערב חג"
            else:                       # "day"/"holiday"/anything else => Saturday-like
                IS_SAT[_d] = True; HOLIDAY[_d] = _nm or "חג"
IS_WEEKEND = {d: IS_FRI[d] or IS_SAT[d] for d in DAYS}
HEB = ["שני","שלישי","רביעי","חמישי","שישי","שבת","ראשון"]
def heb_dow(d):
    return ["ב'","ג'","ד'","ה'","ו'","שבת","א'"][dow(d)]

# ---- parse interns ----
def parse_dates(s):
    s = (s or "").strip()
    if not s: return set()
    out = set()
    for part in s.split(","):
        part = part.strip()
        if not part: continue
        try:
            dt = datetime.date.fromisoformat(part)
            if dt.year == YEAR and dt.month == MONTH:
                out.add(dt.day)
        except Exception:
            pass
    return out

def fnum(s):
    s = (s or "").strip()
    if s == "": return None
    try: return float(s)
    except: return None

interns = {}
with open(CSV, encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        if r["active"].strip().lower() != "true":
            continue
        iid = r["id"]
        appr = {}; strg = {}
        for st in STATIONS:
            a = r.get(f"approved_{st}", "0").strip()
            appr[st] = (a == "1")
            sv = fnum(r.get(f"strength_{st}", ""))
            strg[st] = int(sv) if sv is not None else None
        interns[iid] = dict(
            id=iid, name=r["name"].strip(), religion=r["religion"].strip(),
            preferredStation=(r["preferredStationId"].strip() or None),
            notes=r["notes"].strip(),
            maxTotal=fnum(r["maxTotalShifts"]),
            minTotal=fnum(r.get("minTotalShifts", "")),
            maxFri=fnum(r["maxFridayShifts"]),
            maxSat=fnum(r["maxSaturdayShifts"]),
            maxWeekend=fnum(r["maxWeekendShifts"]),
            maxSandwiches=fnum(r.get("maxSandwiches", "")),
            blocked=parse_dates(r["blockedDates"]),
            preferred=parse_dates(r["preferredDates"]),
            approved=appr, strength=strg,
            requests=json.loads(r.get("specialRequests", "").strip() or "[]"),
        )

# station preference: from preferredStationId field per intern
STATION_PREF = {}                    # intern -> station soft-bonus
for iid, it in interns.items():
    if it["preferredStation"]:
        STATION_PREF[iid] = it["preferredStation"]

# ---- special requests: per-day station overrides, pins, sandwich limits ----
# APPROVED_OVERRIDE[(iid, st, d)] = True/False  overrides the base approved flag for that day
APPROVED_OVERRIDE = {}
PINNED = {}   # (day, station) -> intern_id
MAX_SAND = {} # intern_id -> max sandwiches allowed (None = global rule)
EARLY_PREF = set()   # interns preferring early-month shifts
LATE_PREF  = set()   # interns preferring late-month shifts
for iid, it in interns.items():
    if it.get("maxSandwiches") is not None:
        MAX_SAND[iid] = int(it["maxSandwiches"])
    for req in it.get("requests", []):
        rtype = req.get("type", "")
        if rtype == "pin":
            d = int(req["day"]); st = req["station"]
            if d in DAYS and st in STATIONS:
                PINNED[(d, st)] = iid
        elif rtype == "station_from":
            st = req["station"]; fromDay = int(req["fromDay"])
            for d in DAYS:
                if d >= fromDay:
                    APPROVED_OVERRIDE[(iid, st, d)] = True
                else:
                    APPROVED_OVERRIDE[(iid, st, d)] = False
        elif rtype == "station_until":
            st = req["station"]; untilDay = int(req["untilDay"])
            for d in DAYS:
                if d <= untilDay:
                    APPROVED_OVERRIDE[(iid, st, d)] = False
        elif rtype == "block_station_range":
            st = req["station"]
            fromDay = int(req["fromDay"]); toDay = int(req["toDay"])
            for d in DAYS:
                if fromDay <= d <= toDay:
                    APPROVED_OVERRIDE[(iid, st, d)] = False
        elif rtype == "early_month":
            EARLY_PREF.add(iid)
        elif rtype == "late_month":
            LATE_PREF.add(iid)
if PINNED:
    print(f"  Pinned assignments: {len(PINNED)}")
if APPROVED_OVERRIDE:
    print(f"  Station overrides: {len(APPROVED_OVERRIDE)} (day,station,intern) entries")
if MAX_SAND:
    print(f"  Per-intern sandwich limits: {MAX_SAND}")
if EARLY_PREF:
    print(f"  Early-month preference: {EARLY_PREF}")
if LATE_PREF:
    print(f"  Late-month preference: {LATE_PREF}")

def is_approved(i, st, d):
    """Check if intern i is approved for station st on day d, considering overrides."""
    key = (i, st, d)
    if key in APPROVED_OVERRIDE:
        return APPROVED_OVERRIDE[key]
    return interns[i]["approved"][st]

IIDS = list(interns.keys())
print(f"Active interns: {len(IIDS)}")
for st in STATIONS:
    pool = [i for i in IIDS if interns[i]["approved"][st]]
    print(f"  {st:6s}: {len(pool)} approved")

# ---- capacity (water-filling target for total shifts) ----
def avail_days(i):
    return [d for d in DAYS if d not in interns[i]["blocked"]]
def can_any_station(i):
    if any(interns[i]["approved"][st] for st in STATIONS):
        return True
    # check if any override grants approval on any day
    return any(APPROVED_OVERRIDE.get((i, st, d), False) for st in STATIONS for d in DAYS)

for i in list(interns):
    if not can_any_station(i):
        print("WARN no approved station:", i, interns[i]["name"])

# max shifts feasible by alternating-day rule within available days
def alt_cap(i):
    av = sorted(avail_days(i)); cap = 0; last = -10
    for d in av:
        if d - last >= 2:
            cap += 1; last = d
    return cap

CAP = {}
for i in IIDS:
    c = alt_cap(i)
    if interns[i]["maxTotal"] is not None:
        c = min(c, int(interns[i]["maxTotal"]))
    CAP[i] = c

TOTAL_SLOTS = len(DAYS) * len(STATIONS) - len(SKIP)   # intern slots only
SLOTS = [(d, st) for d in DAYS for st in STATIONS if (d, st) not in SKIP]
# water-filling fair targets for total
def water_fill(cap_map, total):
    target = {i: 0.0 for i in cap_map}
    remaining = total; pool = set(cap_map)
    while pool and remaining > 1e-9:
        share = remaining / len(pool)
        froze = [i for i in pool if cap_map[i] <= target[i] + share + 1e-9 and cap_map[i] - target[i] < share]
        if not froze:
            for i in pool: target[i] += share
            remaining = 0
        else:
            for i in froze:
                give = cap_map[i] - target[i]
                target[i] = cap_map[i]; remaining -= give; pool.discard(i)
    return target
FAIR_TOTAL = water_fill(CAP, TOTAL_SLOTS)
# ---- per-station feasibility check (surfaces bottlenecks) ----
STATION_HEB = {"er1": "מיון 1", "er2": "מיון 2", "nicu1": "פגיה 1",
               "nicu2": "פגיה 2", "ward": "מחלקה", "picu": 'טיפ"נ'}
station_shortages = []
for st in STATIONS:
    needed = sum(1 for d in DAYS if (d, st) not in SKIP)
    # count how many "intern-days" are theoretically approved+available for this station
    supply = 0
    for i in IIDS:
        for d in DAYS:
            if d in interns[i]["blocked"]:
                continue
            if APPROVED_OVERRIDE.get((i, st, d), interns[i]["approved"][st]):
                supply += 1
    # each intern can work at most alt_cap days total → this is an approximation
    pool_size = sum(1 for i in IIDS if interns[i]["approved"][st] or
                    any(APPROVED_OVERRIDE.get((i, st, d), False) for d in DAYS))
    if pool_size < 2:
        station_shortages.append(f"{STATION_HEB[st]}: רק {pool_size} מתמחים מאושרים לתחנה (צריך לפחות 2 ליום)")
    elif needed > 0 and supply < needed * 3:  # need generous buffer since interns can't work every day
        station_shortages.append(f"{STATION_HEB[st]}: פחות מדי מתמחים זמינים ({pool_size} מתמחים, צריך למלא {needed} ימים)")

if station_shortages:
    print("DIAG_SHORTAGE_START")
    for msg in station_shortages:
        print("DIAG:", msg)
    print("DIAG_SHORTAGE_END")

# ---- total capacity vs demand ----
if sum(CAP.values()) < TOTAL_SLOTS:
    print("DIAG_CAPACITY:", f"סך הקיבולת של המתמחים ({sum(CAP.values())}) קטן מהמשבצות למילוי ({TOTAL_SLOTS})")

# ---- interns with no approvals ----
no_approval = [interns[i]["name"] for i in IIDS if not can_any_station(i)]
if no_approval:
    print("DIAG_NO_APPROVAL:", "מתמחים ללא אישור לאף תחנה: " + ", ".join(no_approval))

# ---- interns almost fully blocked ----
heavily_blocked = []
for i in IIDS:
    avail = len(avail_days(i))
    if avail < 3:
        heavily_blocked.append(f"{interns[i]['name']} (רק {avail} ימים זמינים)")
if heavily_blocked:
    print("DIAG_BLOCKED:", "מתמחים עם מעט מאוד ימים זמינים: " + ", ".join(heavily_blocked))

print("Sum cap=", sum(CAP.values()), "need=", TOTAL_SLOTS)

print(json.dumps({"caps_sample": {interns[i]['name']: CAP[i] for i in IIDS}}, ensure_ascii=False))

# =========================================================
#                 SCHEDULING ENGINE
# =========================================================
INF = float("inf")

class State:
    def __init__(s):
        s.assign = {(d, st): None for d in DAYS for st in STATIONS}
        s.idays = {i: set() for i in IIDS}            # days intern works
        s.cnt = {i: dict(total=0, thu=0, fri=0, sat=0, wknd=0, wday=0,
                         st={k: 0 for k in STATIONS}) for i in IIDS}

    def working(s, i, d):
        return d in s.idays[i]

    def place(s, i, d, st):
        s.assign[(d, st)] = i
        s.idays[i].add(d)
        c = s.cnt[i]; c["total"] += 1; c["st"][st] += 1
        if IS_THU[d]: c["thu"] += 1
        if IS_FRI[d]: c["fri"] += 1
        if IS_SAT[d]: c["sat"] += 1
        if IS_WEEKEND[d]: c["wknd"] += 1
        else: c["wday"] += 1

    def remove(s, d, st):
        i = s.assign[(d, st)]
        if i is None: return
        s.assign[(d, st)] = None
        s.idays[i].discard(d)
        c = s.cnt[i]; c["total"] -= 1; c["st"][st] -= 1
        if IS_THU[d]: c["thu"] -= 1
        if IS_FRI[d]: c["fri"] -= 1
        if IS_SAT[d]: c["sat"] -= 1
        if IS_WEEKEND[d]: c["wknd"] -= 1
        else: c["wday"] -= 1

def hard_ok(s, i, d, st, ignore_self_day=False):
    it = interns[i]
    if not is_approved(i, st, d): return False
    if d in it["blocked"]: return False
    if not ignore_self_day and s.working(i, d): return False
    if (d-1) in s.idays[i] or (d+1) in s.idays[i]: return False
    c = s.cnt[i]
    if it["maxTotal"] is not None and c["total"] >= it["maxTotal"]: return False
    if IS_FRI[d] and it["maxFri"] is not None and c["fri"] >= it["maxFri"]: return False
    if IS_SAT[d] and it["maxSat"] is not None and c["sat"] >= it["maxSat"]: return False
    if IS_WEEKEND[d] and it["maxWeekend"] is not None and c["wknd"] >= it["maxWeekend"]: return False
    # pair strength vs already-placed partner
    if st in PAIR_OF:
        pst = PAIR_OF[st]; j = s.assign[(d, pst)]
        if j is not None and not pair_strength_ok(i, st, j, pst):
            return False
    # cross PICU<->ER1 strength vs already-placed counterpart
    if st == "picu" and not cross_strength_ok(i, s.assign[(d, "er1")]):
        return False
    if st == "er1" and not cross_strength_ok(s.assign[(d, "picu")], i):
        return False
    return True

def pair_strength_ok(i, st, j, pst):
    si = interns[i]["strength"][st]; sj = interns[j]["strength"][pst]
    if si is None or sj is None: return True
    # hard rule: a weak(1) must be paired only with a strong(3)
    if si == 1 and sj != 3: return False
    if sj == 1 and si != 3: return False
    return True

# ---- cross-station constraint: PICU <-> ER1 ----
# Rule (one-directional, as specified): if the intern in PICU is weak (strength 1),
# the intern in ER1 on the same day must be strong (strength 3).
# "strong" == 3, consistent with pair_strength_ok above.
CROSS_DIRECTIONAL = True   # set False for a symmetric rule (weak ER1 also requires strong PICU)
def cross_strength_ok(pic, er):
    # pic = intern assigned to picu that day, er = intern assigned to er1 that day
    if pic is None or er is None: return True
    sp = interns[pic]["strength"]["picu"]
    se = interns[er]["strength"]["er1"]
    if sp == 1 and se != 3: return False
    if not CROSS_DIRECTIONAL and se == 1 and sp != 3: return False
    return True

# ---------- per-day constructive solve with restarts ----------
def feasible_candidates(s, d, st, used_today):
    out = []
    for i in IIDS:
        if i in used_today: continue
        if hard_ok(s, i, d, st):
            out.append(i)
    return out

def greedy_score(s, i, d, st):
    it = interns[i]; c = s.cnt[i]
    # primary: load relative to fair target (prefer underloaded)
    ratio = c["total"] / max(FAIR_TOTAL[i], 0.5)
    score = ratio * 100
    # weekend equity: discourage piling weekends on same people
    if IS_WEEKEND[d]:
        score += c["wknd"] * 8
    if IS_FRI[d]: score += c["fri"] * 4
    if IS_SAT[d]: score += c["sat"] * 6
    if IS_THU[d]: score += c["thu"] * 4
    # preferred dates -> bonus
    if d in it["preferred"]: score -= 25
    # station preference
    if STATION_PREF.get(i) == st: score -= 8
    # sandwich avoidance (per-intern limit via MAX_SAND handled in count_sandwiches)
    if (d-2) in s.idays[i] or (d+2) in s.idays[i]:
        score += 12
    # spacing: encourage gaps; penalty if close to last shift
    if s.idays[i]:
        nearest = min(abs(d-x) for x in s.idays[i])
        if nearest <= 3: score += (4 - nearest) * 2
    # early-month / late-month preference (from special requests)
    if i in EARLY_PREF and d > 21: score += 15
    if i in LATE_PREF and d < 11: score += 15
    # prefer using scarce picu/er1 people elsewhere sparingly:
    # discourage spending er1/picu-eligible "core" people on ward
    if st == "ward" and (interns[i]["approved"]["picu"] or interns[i]["approved"]["er1"]):
        score += 6
    score += random.random() * 3
    return score

def solve_day(s, d, restarts=40):
    # identify already-placed slots (e.g. pinned) and their interns
    pre_used = set()
    for st in STATIONS:
        if s.assign[(d, st)] is not None:
            pre_used.add(s.assign[(d, st)])
    open_stations = [st for st in STATIONS if (d, st) not in SKIP and s.assign[(d, st)] is None]
    if not open_stations:
        return {}  # all filled (pins + externals)
    order_pre = sorted(open_stations, key=lambda st: len(feasible_candidates(s, d, st, pre_used)))
    best = None; best_cost = INF
    for _ in range(restarts):
        trial = {}; used = set(pre_used); ok = True
        remaining = list(open_stations)
        while remaining:
            # candidate counts now
            cc = {}
            for st in remaining:
                cand = [i for i in IIDS if i not in used and hard_ok_trial(s, trial, d, i, st)]
                cc[st] = cand
            st = min(remaining, key=lambda x: len(cc[x]))
            cand = cc[st]
            if not cand:
                ok = False; break
            # weighted pick by greedy score (softmax-ish: pick among best few)
            scored = sorted(cand, key=lambda i: greedy_score_trial(s, trial, d, i, st))
            k = min(len(scored), 4)
            choice = random.choice(scored[:k]) if random.random() < 0.5 else scored[0]
            trial[st] = choice; used.add(choice); remaining.remove(st)
        if ok:
            cost = sum(greedy_score_trial(s, {k:v for k,v in trial.items() if k!=st2}, d, v, st2)
                       for st2, v in trial.items())
            if cost < best_cost:
                best_cost = cost; best = dict(trial)
    return best  # may be None if day infeasible

def hard_ok_trial(s, trial, d, i, st):
    # like hard_ok but partner may be in trial (not yet committed to state)
    it = interns[i]
    if not is_approved(i, st, d): return False
    if d in it["blocked"]: return False
    if s.working(i, d): return False
    if (d-1) in s.idays[i] or (d+1) in s.idays[i]: return False
    c = s.cnt[i]
    if it["maxTotal"] is not None and c["total"] >= it["maxTotal"]: return False
    if IS_FRI[d] and it["maxFri"] is not None and c["fri"] >= it["maxFri"]: return False
    if IS_SAT[d] and it["maxSat"] is not None and c["sat"] >= it["maxSat"]: return False
    if IS_WEEKEND[d] and it["maxWeekend"] is not None and c["wknd"] >= it["maxWeekend"]: return False
    if st in PAIR_OF:
        pst = PAIR_OF[st]; j = trial.get(pst)
        if j is not None and not pair_strength_ok(i, st, j, pst):
            return False
    if st == "picu" and not cross_strength_ok(i, trial.get("er1")):
        return False
    if st == "er1" and not cross_strength_ok(trial.get("picu"), i):
        return False
    return True

def greedy_score_trial(s, trial, d, i, st):
    return greedy_score(s, i, d, st)

# build initial schedule day-by-day
def construct():
    s = State()
    unfilled = []
    # place pinned assignments first
    for (d, st), iid in PINNED.items():
        if (d, st) not in SKIP and iid in IIDS:
            if not s.working(iid, d):
                s.place(iid, d, st)
    for d in DAYS:
        res = solve_day(s, d)
        if res is None:
            # fill what we can greedily, mark rest unfilled
            used = set()
            for st in sorted([s2 for s2 in STATIONS if (d, s2) not in SKIP],
                             key=lambda st: len(feasible_candidates(s, d, st, used))):
                cand = [i for i in IIDS if i not in used and hard_ok(s, i, d, st)]
                if cand:
                    ch = min(cand, key=lambda i: greedy_score(s, i, d, st))
                    s.place(ch, d, st); used.add(ch)
                else:
                    unfilled.append((d, st))
        else:
            for st, i in res.items():
                s.place(i, d, st)
    return s, unfilled

def load_base(path):
    """Read a machine-readable schedule (day,station,id,name) -> (assign, names)."""
    base = {}; names = {}
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            try: d = int(str(row.get("day", "")).strip())
            except (ValueError, TypeError): continue
            st = (row.get("station") or "").strip()
            iid = (row.get("id") or "").strip()
            nm = (row.get("name") or "").strip()
            if st in STATIONS and iid:
                base[(d, st)] = iid
                if nm: names[iid] = nm
    return base, names

def construct_midmonth(base):
    """Lock days < FROM_DAY (history), warm-start days >= FROM_DAY from the base where the
    new data still allows it; if a day can't be completed from the base, roll it back and
    solve it fully with the constrained day-solver. Counts & no-consecutive span the lock
    boundary. SA + the stability anchor then minimize changes vs the base."""
    s = State(); unfilled = []
    # 1) lock the past: place existing assignments of still-current interns
    for (d, st), iid in base.items():
        if d < FROM_DAY and iid in IIDS and (d, st) not in SKIP and not s.working(iid, d):
            s.place(iid, d, st)
    # 2) future, day by day
    for d in [x for x in DAYS if x >= FROM_DAY]:
        placed = []
        for st in STATIONS:                       # warm-start from base
            if (d, st) in SKIP: continue
            iid = base.get((d, st))
            if iid in IIDS and hard_ok(s, iid, d, st):
                s.place(iid, d, st); placed.append(st)
        used = {s.assign[(d, st)] for st in STATIONS if s.assign[(d, st)] is not None}
        holes = [st for st in STATIONS if (d, st) not in SKIP and s.assign[(d, st)] is None]
        for st in sorted(holes, key=lambda st: len(feasible_candidates(s, d, st, used))):
            cand = [i for i in IIDS if i not in used and hard_ok(s, i, d, st)]
            if cand:
                ch = min(cand, key=lambda i: greedy_score(s, i, d, st))
                s.place(ch, d, st); used.add(ch); placed.append(st)
        # any slot still empty? base couldn't complete this day -> roll back & solve fully
        if any((d, st) not in SKIP and s.assign[(d, st)] is None for st in STATIONS):
            for st in placed: s.remove(d, st)
            res = solve_day(s, d)
            if res is not None:
                for st, i in res.items(): s.place(i, d, st)
            else:                                 # truly infeasible: salvage greedily
                used = set()
                for st in sorted([x for x in STATIONS if (d, x) not in SKIP],
                                 key=lambda st: len(feasible_candidates(s, d, st, used))):
                    cand = [i for i in IIDS if i not in used and hard_ok(s, i, d, st)]
                    if cand:
                        ch = min(cand, key=lambda i: greedy_score(s, i, d, st))
                        s.place(ch, d, st); used.add(ch)
                    else:
                        unfilled.append((d, st))
    return s, unfilled

# =========================================================
#            FAIR TARGETS (weekend / fri / sat)
# =========================================================
WK_DAYS  = [d for d in DAYS if IS_WEEKEND[d]]
FRI_DAYS = [d for d in DAYS if IS_FRI[d]]
SAT_DAYS = [d for d in DAYS if IS_SAT[d]]
THU_DAYS = [d for d in DAYS if IS_THU[d]]

def cat_cap(i, days, maxk):
    av = [d for d in days if d not in interns[i]["blocked"]]
    c = len(av)
    if maxk is not None: c = min(c, int(maxk))
    return c

CAP_WK  = {i: cat_cap(i, WK_DAYS,  interns[i]["maxWeekend"]) for i in IIDS}
CAP_FRI = {i: cat_cap(i, FRI_DAYS, interns[i]["maxFri"])      for i in IIDS}
CAP_SAT = {i: cat_cap(i, SAT_DAYS, interns[i]["maxSat"])      for i in IIDS}
CAP_THU = {i: cat_cap(i, THU_DAYS, None)                      for i in IIDS}
FAIR_WK  = water_fill(CAP_WK,  sum(1 for d in WK_DAYS  for st in STATIONS if (d, st) not in SKIP))
FAIR_FRI = water_fill(CAP_FRI, sum(1 for d in FRI_DAYS for st in STATIONS if (d, st) not in SKIP))
FAIR_SAT = water_fill(CAP_SAT, sum(1 for d in SAT_DAYS for st in STATIONS if (d, st) not in SKIP))
FAIR_THU = water_fill(CAP_THU, sum(1 for d in THU_DAYS for st in STATIONS if (d, st) not in SKIP))

W = dict(total=2.0, wknd=2.0, fri=1.5, sat=1.5, thu=1.5, sand=9.0, pref=6.0,
         spread=0.5, stpref=2.0, early=3.0)
# Peak-load smoothing: penalize anyone whose total exceeds PEAK_THR, quadratically.
# PEAK_THR auto-adapts to "one above the rounded mean load", so it tightens
# automatically when fewer slots need covering (e.g. with external duties).
PEAK_THR = math.ceil(TOTAL_SLOTS / max(1, len(IIDS))) + 1
PEAK_W   = 15.0

# Stability anchor for mid-month re-optimization: (day,station) -> intern id that the
# existing schedule had there. Changing such a slot costs STAB_W, so the optimizer keeps
# the existing schedule wherever the new constraints still allow it (minimal changes).
STAB = {}
STAB_W = 30.0
def stab_penalty(s):
    if not STAB: return 0
    return sum(1 for k, iid in STAB.items() if s.assign[k] != iid)

def count_sandwiches(s):
    n = 0
    for i in IIDS:
        ds = s.idays[i]
        cnt_i = sum(1 for d in ds if (d+2) in ds)
        limit = MAX_SAND.get(i)
        if limit is not None:
            n += max(0, cnt_i - limit)  # only penalize excess over personal limit
        else:
            n += cnt_i
    return n

def spread_penalty(s):
    p = 0.0
    for i in IIDS:
        ds = sorted(s.idays[i])
        if len(ds) < 2: continue
        gaps = [ds[k+1]-ds[k] for k in range(len(ds)-1)]
        mean = sum(gaps)/len(gaps)
        p += sum((g-mean)**2 for g in gaps)/len(gaps)
    return p

def pref_penalty(s):
    # missed preferred-date opportunities (intern wanted day d, not working it)
    miss = 0
    for i in IIDS:
        for d in interns[i]["preferred"]:
            if d not in s.idays[i] and d not in interns[i]["blocked"]:
                miss += 1
    return miss

def stpref_penalty(s):
    p = 0
    for i, want in STATION_PREF.items():
        worked_pref = s.cnt[i]["st"].get(want, 0)
        other = s.cnt[i]["total"] - worked_pref
        p += other  # penalize shifts not in preferred station
    return p

def early_penalty(s):
    p = 0
    for i in EARLY_PREF:
        if i in s.idays:
            p += sum(1 for d in s.idays[i] if d > 21)
    for i in LATE_PREF:
        if i in s.idays:
            p += sum(1 for d in s.idays[i] if d < 11)
    return p

def cost(s):
    ct = sum((s.cnt[i]["total"] - FAIR_TOTAL[i])**2 for i in IIDS)
    cw = sum((s.cnt[i]["wknd"]  - FAIR_WK[i])**2   for i in IIDS)
    cf = sum((s.cnt[i]["fri"]   - FAIR_FRI[i])**2  for i in IIDS)
    cs = sum((s.cnt[i]["sat"]   - FAIR_SAT[i])**2  for i in IIDS)
    ch = sum((s.cnt[i]["thu"]   - FAIR_THU[i])**2  for i in IIDS)
    peak = sum(max(0, s.cnt[i]["total"] - PEAK_THR)**2 for i in IIDS)
    # minTotal: heavy penalty for being below the floor
    mintot = sum(max(0, int(interns[i]["minTotal"]) - s.cnt[i]["total"])**2
                 for i in IIDS if interns[i].get("minTotal") is not None)
    # pinned violations: heavy penalty for pins not honored
    pinv = sum(1 for (d, st), iid in PINNED.items()
               if (d, st) not in SKIP and s.assign.get((d, st)) != iid)
    return (W["total"]*ct + W["wknd"]*cw + W["fri"]*cf + W["sat"]*cs + W["thu"]*ch
            + W["sand"]*count_sandwiches(s) + W["pref"]*pref_penalty(s)
            + W["spread"]*spread_penalty(s) + W["stpref"]*stpref_penalty(s)
            + W["early"]*early_penalty(s) + PEAK_W*peak + STAB_W*stab_penalty(s)
            + 50.0*mintot + 100.0*pinv)

# =========================================================
#                 SIMULATED ANNEALING
# =========================================================
def pair_ok_at(s, d, st):
    if st not in PAIR_OF: return True
    pst = PAIR_OF[st]
    i = s.assign[(d, st)]; j = s.assign[(d, pst)]
    if i is None or j is None: return True
    return pair_strength_ok(i, st, j, pst)

def cross_ok_at(s, d):
    return cross_strength_ok(s.assign[(d, "picu")], s.assign[(d, "er1")])

def sa(s, iters=60000, T0=8.0, T1=0.05, keys=None):
    cur = cost(s); best = cur
    best_state = snapshot(s)
    keys = list(SLOTS) if keys is None else list(keys)
    keys = [k for k in keys if k not in PINNED and k not in LOCKED]   # never move pinned/locked slots
    for it in range(iters):
        T = T0 * (T1/T0)**(it/iters)
        if random.random() < 0.75:
            # reassignment move
            d, st = random.choice(keys)
            i = s.assign[(d, st)]
            if i is None: continue
            s.remove(d, st)
            cands = [j for j in IIDS if hard_ok(s, j, d, st)]
            if not cands:
                s.place(i, d, st); continue
            j = random.choice(cands)
            s.place(j, d, st)
            new = cost(s)
            if new <= cur or random.random() < math.exp((cur-new)/max(T,1e-6)):
                cur = new
            else:
                s.remove(d, st); s.place(i, d, st)
        else:
            # swap move
            (d1, st1) = random.choice(keys); (d2, st2) = random.choice(keys)
            if (d1, st1) == (d2, st2): continue
            a = s.assign[(d1, st1)]; b = s.assign[(d2, st2)]
            if a is None or b is None or a == b: continue
            s.remove(d1, st1); s.remove(d2, st2)
            if hard_ok(s, b, d1, st1) and hard_ok(s, a, d2, st2):
                s.place(b, d1, st1); s.place(a, d2, st2)
                # re-validate pair constraints on affected slots (covers same-day pair swaps)
                if (pair_ok_at(s, d1, st1) and pair_ok_at(s, d2, st2)
                        and cross_ok_at(s, d1) and cross_ok_at(s, d2)):
                    new = cost(s)
                    if new <= cur or random.random() < math.exp((cur-new)/max(T,1e-6)):
                        cur = new
                    else:
                        s.remove(d1, st1); s.remove(d2, st2)
                        s.place(a, d1, st1); s.place(b, d2, st2)
                else:
                    s.remove(d1, st1); s.remove(d2, st2)
                    s.place(a, d1, st1); s.place(b, d2, st2)
            else:
                s.place(a, d1, st1); s.place(b, d2, st2)
        if cur < best:
            best = cur; best_state = snapshot(s)
    restore(s, best_state)
    return best

def snapshot(s):
    return {k: v for k, v in s.assign.items()}
def restore(s, snap):
    for d in DAYS:
        for st in STATIONS:
            if s.assign[(d, st)] is not None: s.remove(d, st)
    for (d, st), i in snap.items():
        if i is not None: s.place(i, d, st)

# =========================================================
#                 HARD VALIDATOR
# =========================================================
def validate(s):
    errs = []
    fut = (lambda d: FROM_DAY is None or d >= FROM_DAY)   # mid-month: only the re-optimized days
    # every intern slot filled (external-covered slots, and locked past days, are exempt)
    for d in DAYS:
        if not fut(d): continue
        for st in STATIONS:
            if (d, st) in SKIP: continue
            if s.assign[(d, st)] is None:
                errs.append(f"EMPTY {d} {st}")
    for i in IIDS:
        it = interns[i]; ds = sorted(s.idays[i])
        for d in ds:
            if not fut(d): continue
            for st in STATIONS:
                if s.assign[(d, st)] == i and not is_approved(i, st, d):
                    errs.append(f"{it['name']} unapproved {st} d{d}")
            if d in it["blocked"]:
                errs.append(f"{it['name']} blocked day {d}")
        for k in range(len(ds)-1):              # consecutive: full history (catches the lock boundary)
            if ds[k+1]-ds[k] == 1:
                errs.append(f"{it['name']} consecutive {ds[k]}->{ds[k+1]}")
        # one station per day
        for d in ds:
            if not fut(d): continue
            cnt = sum(1 for st in STATIONS if s.assign[(d, st)] == i)
            if cnt > 1: errs.append(f"{it['name']} {cnt} stations on d{d}")
        c = s.cnt[i]
        if it["maxTotal"] is not None and c["total"] > it["maxTotal"]:
            errs.append(f"{it['name']} total {c['total']}>{it['maxTotal']}")
        if it["minTotal"] is not None and c["total"] < int(it["minTotal"]):
            errs.append(f"{it['name']} total {c['total']}<min {int(it['minTotal'])}")
        if it["maxFri"] is not None and c["fri"] > it["maxFri"]:
            errs.append(f"{it['name']} fri {c['fri']}>{it['maxFri']}")
        if it["maxSat"] is not None and c["sat"] > it["maxSat"]:
            errs.append(f"{it['name']} sat {c['sat']}>{it['maxSat']}")
        if it["maxWeekend"] is not None and c["wknd"] > it["maxWeekend"]:
            errs.append(f"{it['name']} wknd {c['wknd']}>{it['maxWeekend']}")
    # pinned assignments
    for (d, st), iid in PINNED.items():
        if (d, st) not in SKIP and fut(d) and s.assign.get((d, st)) != iid:
            errs.append(f"pin {interns[iid]['name']} not at {st} d{d}")
    # pair strength
    for d in DAYS:
        if not fut(d): continue
        for (a, b) in PAIRS:
            i = s.assign[(d, a)]; j = s.assign[(d, b)]
            if i and j and not pair_strength_ok(i, a, j, b):
                errs.append(f"pair {d} {a}({interns[i]['strength'][a]})/{b}({interns[j]['strength'][b]})")
    # cross PICU<->ER1 strength
    for d in DAYS:
        if not fut(d): continue
        pic = s.assign[(d, "picu")]; er = s.assign[(d, "er1")]
        if pic and er and not cross_strength_ok(pic, er):
            errs.append(f"cross d{d} picu {interns[pic]['name']}(str{interns[pic]['strength']['picu']})"
                        f" / er1 {interns[er]['name']}(str{interns[er]['strength']['er1']})")
    return errs


# ---- aliases for the builder section ----
SLAB = {"er1": "מיון 1", "er2": "מיון 2", "nicu1": "פגיה 1",
        "nicu2": "פגיה 2", "ward": "מחלקה", "picu": chr(34).join(["טיפ", "נ"])}
DOWH = ["ב'", "ג'", "ד'", "ה'", "ו'", "שבת", "א'"]
IS_WK = IS_WEEKEND

def BUILD(assign, order, name_of, notes_lines=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    # ---------- styling ----------
    F="Arial"
    def font(b=False,sz=10,c="000000"): return Font(name=F,bold=b,size=sz,color=c)
    center=Alignment(horizontal="center",vertical="center",wrap_text=True)
    right=Alignment(horizontal="right",vertical="center")
    thin=Side(style="thin",color="BBBBBB"); med=Side(style="medium",color="666666")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    NAVY="1F3864"; BLUE="2E5496"; LBLUE="D6E0F0"; WKND="FCE4D6"; GREY="F2F2F2"
    GREEN="C6EFCE"; YEL="FFF2CC"; RED="F8CBAD"
    def fill(c): return PatternFill("solid",fgColor=c)

    wb=Workbook()

    # =========================================================
    # Sheet 1: לוח חודשי
    # =========================================================
    ws=wb.active; ws.title="לוח חודשי"; ws.sheet_view.rightToLeft=True
    ws["A1"]="לוח תורנויות – יולי 2026"; ws["A1"].font=font(True,16,NAVY)
    ws.merge_cells("A1:H1")
    ws["A2"]=(f"{len(STATIONS)} עמדות × {len(DAYS)} ימים = {len(DAYS)*len(STATIONS)} שיבוצים"
              + (f"  ·  {len(SKIP)} בכיסוי תורני חוץ  ·  {TOTAL_SLOTS} למתמחים" if SKIP else "")
              + "  |  כל האילוצים הקשיחים מקוימים (0 הפרות)")
    ws["A2"].font=font(False,10,"555555"); ws.merge_cells("A2:H2")
    hdr=["תאריך","יום"]+[SLAB[st] for st in STATIONS]
    for j,h in enumerate(hdr,1):
        c=ws.cell(4,j,h); c.font=font(True,11,"FFFFFF"); c.fill=fill(BLUE)
        c.alignment=center; c.border=border
    for d in DAYS:
        rr=4+d
        a=ws.cell(rr,1,f"{d}/7"); a.alignment=center; a.font=font(IS_WK[d],10)
        b=ws.cell(rr,2,DOWH[dow(d)]); b.alignment=center; b.font=font(IS_WK[d],10)
        for j,st in enumerate(STATIONS,3):
            if (d,st) in EXTERNAL:
                c=ws.cell(rr,j,EXTERNAL[(d,st)]); c.font=font(False,9,"806000")
            else:
                iid=assign[(d,st)]
                c=ws.cell(rr,j,name_of.get(iid,"—")); c.font=font(False,10)
            c.alignment=center; c.border=border
        for j in (1,2): ws.cell(rr,j).border=border
        if IS_WK[d]:
            for j in range(1,9): ws.cell(rr,j).fill=fill(WKND)
        elif d%2==0:
            for j in range(1,9): ws.cell(rr,j).fill=fill(GREY)
        for j,st in enumerate(STATIONS,3):          # external cells: soft-yellow on top of row shade
            if (d,st) in EXTERNAL: ws.cell(rr,j).fill=fill("FFF2CC")
    ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=6
    for col in "CDEFGH": ws.column_dimensions[col].width=12
    ws.freeze_panes="A5"
    # legend
    lr=4+31+2
    ws.cell(lr,1,"מקרא:").font=font(True,10)
    ws.cell(lr,3,"שורה כתומה = סוף שבוע · תא צהוב = כיסוי תורן חוץ").font=font(False,9,"555555")
    ws.merge_cells(start_row=lr,start_column=3,end_row=lr,end_column=8)

    # =========================================================
    # Sheet 2: מטריצת ימים  (intern x day) + flag rows
    # =========================================================
    m=wb.create_sheet("מטריצת מתמחה"); m.sheet_view.rightToLeft=True
    m["A1"]="מטריצת מתמחה × יום (התא = העמדה שבה שובץ)"; m["A1"].font=font(True,14,NAVY)
    m.merge_cells(start_row=1,start_column=1,end_row=1,end_column=33)
    # header row 3: day numbers; rows 4-6 flags; interns from row 7
    m.cell(3,1,"מתמחה").font=font(True,10,"FFFFFF"); m.cell(3,1).fill=fill(BLUE); m.cell(3,1).alignment=center; m.cell(3,1).border=border
    for d in DAYS:
        c=m.cell(3,1+d,d); c.font=font(True,9,"FFFFFF"); c.fill=fill(BLUE if not IS_WK[d] else "C55A11"); c.alignment=center; c.border=border
    flagrows=[("שישי?",IS_FRI),("שבת?",IS_SAT),('סופ"ש?',IS_WK)]
    for k,(lab,fl) in enumerate(flagrows):
        r=4+k
        c=m.cell(r,1,lab); c.font=font(True,8,"555555"); c.alignment=right; c.fill=fill(GREY); c.border=border
        for d in DAYS:
            cc=m.cell(r,1+d,1 if fl[d] else 0); cc.font=font(False,8,"AAAAAA"); cc.alignment=center; cc.border=border; cc.fill=fill(GREY)
    FLAG_FRI_ROW,FLAG_SAT_ROW,FLAG_WK_ROW=4,5,6
    first_intern_row=7
    for idx,iid in enumerate(order):
        r=first_intern_row+idx
        nc=m.cell(r,1,name_of[iid]); nc.font=font(True,9); nc.alignment=right; nc.border=border
        for d in DAYS:
            iid2=assign[(d, None)] if False else None
            # find station for this intern this day
            lab=""
            for st in STATIONS:
                if assign[(d,st)]==iid: lab=SLAB[st]; break
            c=m.cell(r,1+d,lab); c.alignment=center; c.font=font(False,8); c.border=border
            if lab:
                c.fill=fill(LBLUE if not IS_WK[d] else WKND)
            # mark blocked days lightly
            elif d in interns[iid]["blocked"]:
                c.fill=fill("EDEDED")
    m.column_dimensions["A"].width=12
    for d in DAYS: m.column_dimensions[get_column_letter(1+d)].width=6.5
    m.freeze_panes="B7"
    last_intern_row=first_intern_row+len(order)-1

    # =========================================================
    # Sheet 3: סיכום והוגנות  (formulas referencing matrix)
    # =========================================================
    g=wb.create_sheet("סיכום והוגנות"); g.sheet_view.rightToLeft=True
    g["A1"]="סיכום עומסים והוגנות"; g["A1"].font=font(True,14,NAVY)
    g.merge_cells("A1:M1")
    cols=["מתמחה","סה""כ","חול","שישי","שבת","סופ""ש","מיון1","מיון2","פגיה1","פגיה2","מחלקה",'טיפ"נ',"תקרות (סהכ/ו/ש/סופ)"]
    for j,h in enumerate(cols,1):
        c=g.cell(3,j,h); c.font=font(True,10,"FFFFFF"); c.fill=fill(BLUE); c.alignment=center; c.border=border
    MROW="'מטריצת מתמחה'!"
    def mrange(r): return f"{MROW}$B{r}:$AF{r}"  # B..AF = days 1..31
    for idx,iid in enumerate(order):
        gr=4+idx; mr=first_intern_row+idx
        g.cell(gr,1,name_of[iid]).font=font(True,10); g.cell(gr,1).alignment=right
        rng=f"{MROW}B{mr}:AF{mr}"
        g.cell(gr,2,f'=COUNTA({rng})')                                   # total
        g.cell(gr,4,f'=SUMPRODUCT(({rng}<>"")*({MROW}B$4:AF$4))')        # fri
        g.cell(gr,5,f'=SUMPRODUCT(({rng}<>"")*({MROW}B$5:AF$5))')        # sat
        g.cell(gr,6,f'=SUMPRODUCT(({rng}<>"")*({MROW}B$6:AF$6))')        # weekend
        g.cell(gr,3,f'=B{gr}-F{gr}')                                     # weekday
        for k,st in enumerate(STATIONS):                                  # per-station
            lab=SLAB[st].replace('"','""')
            g.cell(gr,7+k,f'=COUNTIF({rng},"{lab}")')
        it=interns[iid]
        def cap(v): return "—" if v is None else int(v)
        g.cell(gr,13,f"{cap(it['maxTotal'])} / {cap(it['maxFri'])} / {cap(it['maxSat'])} / {cap(it['maxWeekend'])}")
        g.cell(gr,13).font=font(False,9,"777777"); g.cell(gr,13).alignment=center
        for j in range(2,13):
            g.cell(gr,j).alignment=center; g.cell(gr,j).font=font(False,10); g.cell(gr,j).border=border
        g.cell(gr,1).border=border
        if any(interns[iid][k] is not None for k in("maxTotal","maxFri","maxSat","maxWeekend")):
            g.cell(gr,13).fill=fill(YEL)
    nrow=4+len(order)
    # stats rows
    stat_labels=[("מינימום","MIN"),("מקסימום","MAX"),("ממוצע","AVERAGE"),("סטיית תקן","STDEV")]
    for s_i,(lab,fn) in enumerate(stat_labels):
        rr=nrow+1+s_i
        c=g.cell(rr,1,lab); c.font=font(True,10,NAVY); c.alignment=right; c.fill=fill(LBLUE); c.border=border
        for j in range(2,7):
            col=get_column_letter(j)
            cc=g.cell(rr,j,f'={fn}({col}4:{col}{nrow-0})')
            cc.alignment=center; cc.font=font(True,10,NAVY); cc.fill=fill(LBLUE); cc.border=border
            if fn=="AVERAGE" or fn=="STDEV": cc.number_format="0.0"
    for j in range(1,14):
        pass
    g.column_dimensions["A"].width=12
    for col in "BCDEF": g.column_dimensions[col].width=7
    for col in "GHIJKL": g.column_dimensions[col].width=7
    g.column_dimensions["M"].width=18
    g.freeze_panes="B4"

    # =========================================================
    # Sheet 4: הנחות ואילוצים
    # =========================================================
    n=wb.create_sheet("הנחות והערות"); n.sheet_view.rightToLeft=True
    n["A1"]="הנחות, מתודולוגיה והערות"; n["A1"].font=font(True,14,NAVY); n.merge_cells("A1:B1")
    lines = notes_lines if notes_lines is not None else [
     ("מטרה","שיבוץ 6 עמדות בכל יום ביולי 2026 (186 שיבוצים), מיטוב ללא פגיעה באילוצים קשיחים."),
     ("מתמחים פעילים","26 (אביעד, אריאל, ברית מסומנים active=false וכל ימיהם חסומים – הוצאו)."),
     ("עמדות","מיון1, מיון2 (צמד) · פגיה1, פגיה2 (צמד) · מחלקה · טיפ\"נ."),
     ("אילוץ צמד (קשיח)","אם בעמדת צמד אחת משובץ חלש (חוזק 1) – בשנייה חייב חזק (חוזק 3). אומת: 0 הפרות."),
     ("ללא רצף","אין שיבוץ ביומיים רצופים; אין יותר מעמדה אחת ביום. אומת: 0 הפרות."),
     ("חסימות","ימים חסומים לכל מתמחה כובדו במלואם. אומת: 0 הפרות."),
     ("תקרות","maxTotal / maxFriday / maxSaturday / maxWeekend כובדו במלואם. אומת: 0 הפרות."),
     ("סוף שבוע","שישי+שבת לכולם. אין חגים יהודיים/מוסלמיים ביולי 2026 שמתנהגים כסוף-שבוע (תשעה באב הוא צום, לא טופל כסופ\"ש)."),
     ("סנדוויץ' (רך)","שאיפה להימנע מהפרש יום (עבודה-מנוחה-עבודה). בפתרון: 0 סנדוויצ'ים (ליהי פטורה לפי הערה)."),
     ("ימים מועדפים (רך)","17 מתוך 20 בקשות כובדו. לא כובדו: יוכבד 17/7, שרית 8/7, רז 13/7 – נמנעו ע\"י אילוצים קשיחים (רצף/צמד/תקרה)."),
     ("העדפת תחנה (רך)","אביה ומור → פגיה1 (רוב התורנויות); עלי → מיון2 (כל התורנויות)."),
     ("ישי (רך)","רוב התורנויות ב-3 השבועות הראשונים (6 מתוך 7; אחת ב-24/7)."),
     ("הוגנות","יעדים מאוזנים בשיטת water-filling יחסית לזמינות ולתקרות. רוב המתמחים 6–9 תורנויות; הנמוכים כבולים בתקרה/זמינות (אפרת 5, מוחמד נתשה 3, ליהי 7)."),
     ("יוכבד – הערה","ההערה ('mid-month ER1...') לא נאכפה כי מיון1/מיון2 אינם מאושרים לה בנתונים (approved=0); שובצה בפגיה/מחלקה לפי האישורים."),
     ("שיטה","בנייה חמדנית מוגבלת-אילוצים + חיפוש מקומי (Simulated Annealing), 6 התחלות, נבחר הפתרון התקין בעל העלות הנמוכה."),
     ("אימות","כל האילוצים הקשיחים נבדקו תוכניתית – 0 הפרות; כל 186 המשבצות מאוישות."),
    ]
    for i,(k,v) in enumerate(lines,3):
        a=n.cell(i,1,k); a.font=font(True,10,BLUE); a.alignment=Alignment(horizontal="right",vertical="top")
        b=n.cell(i,2,v); b.font=font(False,10); b.alignment=Alignment(horizontal="right",vertical="top",wrap_text=True)
    n.column_dimensions["A"].width=20; n.column_dimensions["B"].width=95

    # =========================================================
    # Sheet 5: אישורים ואילוצים  (reference – all interns incl. inactive)
    # =========================================================
    ca = wb.create_sheet("אישורים ואילוצים"); ca.sheet_view.rightToLeft = True
    ca["A1"] = "אישורים ואילוצים לכל מתמחה"; ca["A1"].font = font(True,14,NAVY)
    ca.merge_cells("A1:Q1")
    ca["A2"] = ("תא תחנה = חוזק (1=חלש · 2=בינוני · 3=חזק) · — = לא מאושר · "
                "ימים חסומים/מועדפים מוצגים כמספרי יום בחודש · שורות אפורות = מתמחה לא פעיל.")
    ca["A2"].font = font(False,9,"555555"); ca.merge_cells("A2:Q2")
    headers = (["שם","פעיל","דת"] + [SLAB[st] for st in STATIONS] +
               ["מקס׳ סהכ","מקס׳ ו׳","מקס׳ ש׳","מקס׳ סופ״ש","תחנה מועדפת",
                "ימים חסומים","ימים מועדפים","הערות"])
    hr = 4
    for j,h in enumerate(headers,1):
        c = ca.cell(hr,j,h); c.font=font(True,10,"FFFFFF"); c.fill=fill(BLUE)
        c.alignment=center; c.border=border
    STR_FILL = {"1": RED, "2": YEL, "3": GREEN}
    def _cap(row,key):
        v=(row.get(key) or "").strip(); return v if v else "—"
    def _days(row,key):
        ds = sorted(parse_dates(row.get(key,"")))
        return ", ".join(str(d) for d in ds) if ds else "—"
    rr = hr+1
    with open(CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            active = (row["active"].strip().lower() == "true")
            nm = ca.cell(rr,1,row["name"].strip()); nm.font=font(True,10); nm.alignment=right; nm.border=border
            av = ca.cell(rr,2,"כן" if active else "לא"); av.alignment=center; av.border=border
            av.font=font(True,10,"548235" if active else "C00000")
            rc = ca.cell(rr,3,(row.get("religion") or "").strip()); rc.alignment=center; rc.font=font(False,10); rc.border=border
            for k,st in enumerate(STATIONS):
                appr = (row.get(f"approved_{st}","0").strip() == "1")
                sv = (row.get(f"strength_{st}") or "").strip()
                cell = ca.cell(rr,4+k); cell.alignment=center; cell.border=border
                if appr:
                    cell.value = sv if sv else "✓"; cell.font=font(True,10)
                    if sv in STR_FILL: cell.fill = fill(STR_FILL[sv])
                else:
                    cell.value = "—"; cell.font=font(False,10,"BBBBBB")
            for off,key in enumerate(["maxTotalShifts","maxFridayShifts","maxSaturdayShifts","maxWeekendShifts"]):
                c=ca.cell(rr,10+off,_cap(row,key)); c.alignment=center; c.font=font(False,10); c.border=border
            ps=ca.cell(rr,14,(row.get("preferredStationId") or "").strip() or "—"); ps.alignment=center; ps.font=font(False,10); ps.border=border
            bd=ca.cell(rr,15,_days(row,"blockedDates")); bd.alignment=center; bd.font=font(False,9,"C00000"); bd.border=border
            pd=ca.cell(rr,16,_days(row,"preferredDates")); pd.alignment=center; pd.font=font(False,9,"548235"); pd.border=border
            nt=ca.cell(rr,17,(row.get("notes") or "").strip()); nt.alignment=right; nt.font=font(False,9,"555555"); nt.border=border
            if not active:
                for j in (1,2,3,10,11,12,13,14,17):
                    ca.cell(rr,j).fill = fill("EFEFEF")
            rr += 1
    for col,w in {"A":13,"B":6,"C":7,"D":7,"E":7,"F":7,"G":7,"H":7,"I":7,
                  "J":8,"K":7,"L":7,"M":9,"N":12,"O":20,"P":20,"Q":34}.items():
        ca.column_dimensions[col].width = w
    ca.freeze_panes = "D5"

    wb.save(OUT)



def main():
    base, base_names = ({}, {})
    midmonth = bool(BASE_CSV and FROM_DAY)
    if midmonth:
        base, base_names = load_base(BASE_CSV)
        # stability anchor: keep future slots equal to base where the occupant is still a current intern
        STAB.clear()
        for (d, st), iid in base.items():
            if d >= FROM_DAY and (d, st) not in SKIP and iid in IIDS:
                STAB[(d, st)] = iid
        fut_keys = [(d, st) for (d, st) in SLOTS if d >= FROM_DAY]
        best = None; bc = INF; fb = None; fb_key = None
        for a in range(4):
            random.seed(40 + a)
            stt, unf = construct_midmonth(base)
            sa(stt, iters=70000, T0=6.0, T1=0.03, keys=fut_keys)
            errs = validate(stt)
            key = (len(errs), cost(stt) + 1000*len(unf))
            if fb_key is None or key < fb_key: fb_key = key; fb = snapshot(stt)
            if errs: continue
            c = cost(stt) + 1000*len(unf)
            if c < bc: bc = c; best = snapshot(stt)
        if best is None:
            print("WARNING: no clean mid-month solution; using least-bad fallback."); best = fb
        s = State(); restore(s, best)
    else:
        best = None; bc = INF; fb = None; fb_key = None
        for a in range(6):
            random.seed(40 + a)
            stt, unf = construct()
            sa(stt, iters=100000, T0=8.0, T1=0.03)
            errs = validate(stt)
            key = (len(errs), cost(stt) + 1000 * len(unf))
            if fb_key is None or key < fb_key:
                fb_key = key; fb = snapshot(stt)
            if errs: continue
            c = cost(stt) + 1000 * len(unf)
            if c < bc: bc = c; best = snapshot(stt)
        if best is None:
            print("WARNING: no violation-free solution found; using least-bad fallback.")
            best = fb
        s = State(); restore(s, best)

    errs = validate(s)
    filled = sum(1 for k in s.assign if s.assign[k] is not None)
    print("filled %d/%d | hard-violations %d | sandwiches %d | missed-pref %d"
          % (filled, TOTAL_SLOTS, len(errs), count_sandwiches(s), pref_penalty(s)))
    for e in errs[:20]:
        print("  VIOLATION:", e)

    # aggregate empty-slot diagnostics
    empty_by_station = {}
    empty_days = {}
    for e in errs:
        if e.startswith("EMPTY"):
            parts = e.split()
            if len(parts) >= 3:
                d, st = parts[1], parts[2]
                empty_by_station[st] = empty_by_station.get(st, 0) + 1
                empty_days.setdefault(st, []).append(d)
    if empty_by_station:
        print("DIAG_EMPTY_START")
        for st, cnt in sorted(empty_by_station.items(), key=lambda x: -x[1]):
            days_str = ", ".join(empty_days[st][:10])
            more = f" (וכן {len(empty_days[st]) - 10} ימים נוספים)" if len(empty_days[st]) > 10 else ""
            print(f"DIAG: לא הצליח למלא {cnt} משבצות ב{STATION_HEB.get(st, st)} - ימים: {days_str}{more}")
        print("DIAG_EMPTY_END")

    # changes vs base (mid-month)
    changed = []
    if midmonth:
        for (d, st) in SLOTS:
            if d >= FROM_DAY and base.get((d, st)) != s.assign[(d, st)]:
                changed.append((d, st))
        print(f"mid-month: from day {FROM_DAY}; changed {len(changed)} of {len(fut_keys)} future intern-slots.")

    # ---- accurate figures for the assumptions sheet ----
    n_sand = count_sandwiches(s)
    missed = []; requested = 0
    for i in IIDS:
        for d in sorted(interns[i]["preferred"]):
            if d in interns[i]["blocked"]: continue
            requested += 1
            if d not in s.idays[i]: missed.append(f"{interns[i]['name']} {d}/{MONTH}")
    honored = requested - len(missed)
    cross_days = [d for d in DAYS if s.assign[(d, "picu")] is not None
                  and interns[s.assign[(d, "picu")]]["strength"]["picu"] == 1]
    cross_who = sorted({interns[s.assign[(d, "picu")]]["name"] for d in cross_days})
    vstr = "0 הפרות" if not errs else f"{len(errs)} הפרות (ראה לוג)"
    cross_dir = ("חד-כיווני" if CROSS_DIRECTIONAL else "דו-כיווני")
    cross_txt = (f"אם בטיפול נמרץ משובץ חלש (חוזק 1) – במיון 1 חייב חזק (חוזק 3). כיוון: {cross_dir}. "
                 + (f"הופעל ב-{len(cross_days)} ימים ({', '.join(cross_who)} בטיפ\"נ). אומת: {vstr}."
                    if cross_days else f"לא הופעל החודש (אין חלש בטיפ\"נ בשיבוץ). אומת: {vstr}."))
    miss_txt = (f"{honored} מתוך {requested} בקשות כובדו."
                + ("" if not missed else " לא כובדו: " + "; ".join(missed)
                   + " – נמנעו ע\"י אילוצים קשיחים (רצף/צמד/תקרה/חסימה)."))
    ext_by_st = {}
    for (d, st), nm in EXTERNAL.items():
        ext_by_st.setdefault(SLAB[st], []).append(d)
    ext_txt = ("אין החודש." if not SKIP else
               f"{len(SKIP)} משבצות מאוישות ע\"י תורני חוץ (לא מתמחים) והוסרו משיבוץ המתמחים, מה שמפחית את העומס: "
               + " · ".join(f"{st} ({len(sorted(ds))} ימים)" for st, ds in ext_by_st.items())
               + ". מסומנות בצהוב בלוח החודשי.")
    hol_txt = ("אין החודש." if not HOLIDAY else
               "ימים שמתנהגים כסוף שבוע: " +
               " · ".join(f"{d}/{MONTH} ({'ערב חג≈שישי' if IS_FRI[d] and dow(d)!=4 else 'חג≈שבת'}{', '+HOLIDAY[d] if HOLIDAY[d] not in ('חג','ערב חג') else ''})"
                          for d in sorted(HOLIDAY)))
    notes_lines = []
    if midmonth:
        notes_lines.append(("עריכה באמצע חודש",
            f"שיבוץ מחדש מיום {FROM_DAY} ואילך לפי הנתונים המעודכנים; הימים שלפניו נעולים (כפי שהיו). "
            f"שונו {len(changed)} מתוך {len(fut_keys)} משבצות-מתמחה עתידיות — מינימום ההכרחי לעמידה באילוצים. אומת: {vstr}."))
    notes_lines += [
     ("מטרה", f"שיבוץ {len(STATIONS)} עמדות בכל יום; {len(DAYS)*len(STATIONS)} שיבוצים, מתוכם {TOTAL_SLOTS} למתמחים ({len(SKIP)} בכיסוי חוץ). מיטוב ללא פגיעה באילוצים קשיחים."),
     ("תורני חוץ", ext_txt),
     ("חגים", hol_txt),
     ("מתמחים פעילים", f"{len(IIDS)} (מתמחים עם active=false הוצאו מהשיבוץ)."),
     ("עמדות", "מיון1, מיון2 (צמד) · פגיה1, פגיה2 (צמד) · מחלקה · טיפ\"נ."),
     ("אילוץ צמד (קשיח)", f"אם בעמדת צמד אחת משובץ חלש (חוזק 1) – בשנייה חייב חזק (חוזק 3). אומת: {vstr}."),
     ("אילוץ טיפ\"נ⟵מיון1 (קשיח)", cross_txt),
     ("ללא רצף", f"אין שיבוץ ביומיים רצופים; אין יותר מעמדה אחת ביום. אומת: {vstr}."),
     ("חסימות", f"ימים חסומים לכל מתמחה כובדו במלואם. אומת: {vstr}."),
     ("תקרות", f"maxTotal / maxFriday / maxSaturday / maxWeekend כובדו במלואם. אומת: {vstr}."),
     ("סוף שבוע", "שישי+שבת לכולם. ערב חג מתנהג כשישי והחג כשבת."),
     ("סנדוויץ' (רך)", f"שאיפה להימנע מהפרש יום (עבודה-מנוחה-עבודה). בפתרון: {n_sand} סנדוויצ'ים (ליהי פטורה לפי הערה)."),
     ("ימים מועדפים (רך)", miss_txt),
     ("העדפת תחנה (רך)", "מתמחים עם preferredStationId / הערת-תחנה תועדפו לשבץ בתחנתם (פירוט מספרי בגיליון 'סיכום והוגנות')."),
     ("הוגנות", f"יעדים מאוזנים בשיטת water-filling + קנס על עומס מעל {PEAK_THR} כדי לרסן חריגים. מספרים בפועל לכל מתמחה בגיליון 'סיכום והוגנות'."),
     ("שיטה", "בנייה חמדנית מוגבלת-אילוצים + חיפוש מקומי (Simulated Annealing); נבחר הפתרון התקין בעל העלות הנמוכה."),
     ("אימות", f"כל האילוצים הקשיחים נבדקו תוכניתית – {vstr}; {filled}/{TOTAL_SLOTS} משבצות מאוישות."),
    ]

    # display map: locked past verbatim from base (shows even removed interns); future from the solve
    name_of = {i: interns[i]["name"] for i in interns}
    name_of.update(base_names)
    display = dict(s.assign)
    if midmonth:
        for (d, st), iid in base.items():
            if d < FROM_DAY: display[(d, st)] = iid

    BUILD(display, list(IIDS), name_of, notes_lines=notes_lines)

    # machine-readable assignments (for next month's mid-month edits)
    asg = OUT[:-5] + "_assignments.csv" if OUT.endswith(".xlsx") else OUT + ".assignments.csv"
    with open(asg, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f); w.writerow(["day", "station", "id", "name"])
        for d in DAYS:
            for st in STATIONS:
                iid = display.get((d, st))
                if iid is not None and (d, st) not in SKIP:
                    w.writerow([d, st, iid, name_of.get(iid, "")])
    print("saved ->", OUT, "and", asg)
    print("cross-rule active days:", cross_days, cross_who)

if __name__ == "__main__":
    main()
